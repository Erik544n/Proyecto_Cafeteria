from fastapi import APIRouter, Depends, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, timedelta
from io import BytesIO
from pydantic import BaseModel
from typing import Optional
from decimal import Decimal

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# Charts
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np

from ..database import get_db
from ..models import Venta, Pedido, DetallePedido, ProductoMenu, Gasto, Insumo, CategoriaGasto
from .auth import get_current_user, require_rol
from ..models import Usuario

router = APIRouter(prefix="/reportes", tags=["Reportes"])


class ExportarReporteRequest(BaseModel):
    fechaInicio: Optional[str] = None
    fechaFin: Optional[str] = None
    categoriaId: Optional[str] = None
    metodoPago: Optional[str] = None
    formato: str = "pdf"
    incluirProductos: bool = True
    incluirNotas: bool = False
    incluirImpuestos: bool = True
    incluirGraficas: bool = False


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────
def get_ventas_filtradas(db: Session, req: ExportarReporteRequest):
    query = db.query(Venta).join(Pedido, Venta.pedido_id == Pedido.pedido_id).filter(Venta.anulada == False)
    if req.fechaInicio:
        query = query.filter(func.date(Venta.creado_en) >= req.fechaInicio)
    if req.fechaFin:
        query = query.filter(func.date(Venta.creado_en) <= req.fechaFin)
    if req.metodoPago:
        query = query.filter(Venta.metodo_pago == req.metodoPago.upper())
    if req.categoriaId:
        query = query.join(DetallePedido, DetallePedido.pedido_id == Pedido.pedido_id) \
                     .join(ProductoMenu, ProductoMenu.producto_id == DetallePedido.producto_id) \
                     .filter(ProductoMenu.categoria_id == req.categoriaId)
    return query.distinct().all()

def get_productos_filtrados(db: Session, req: ExportarReporteRequest):
    query = db.query(
        ProductoMenu.nombre,
        func.sum(DetallePedido.cantidad).label("total_vendido"),
        func.sum(DetallePedido.subtotal).label("ingresos")
    ).join(
        DetallePedido, ProductoMenu.producto_id == DetallePedido.producto_id
    ).join(
        Pedido, DetallePedido.pedido_id == Pedido.pedido_id
    ).join(
        Venta, Venta.pedido_id == Pedido.pedido_id
    ).filter(
        Pedido.estado.notin_(["CANCELADO"]),
        Venta.anulada == False
    )
    if req.fechaInicio:
        query = query.filter(func.date(Pedido.creado_en) >= req.fechaInicio)
    if req.fechaFin:
        query = query.filter(func.date(Pedido.creado_en) <= req.fechaFin)
    if req.categoriaId:
        query = query.filter(ProductoMenu.categoria_id == req.categoriaId)
    if req.metodoPago:
        query = query.filter(Venta.metodo_pago == req.metodoPago.upper())
    return query.group_by(ProductoMenu.nombre).order_by(func.sum(DetallePedido.cantidad).desc()).all()

def get_gastos_filtrados(db: Session, req: ExportarReporteRequest):
    query = db.query(Gasto)
    if req.fechaInicio:
        query = query.filter(Gasto.fecha_gasto >= req.fechaInicio)
    if req.fechaFin:
        query = query.filter(Gasto.fecha_gasto <= req.fechaFin)
    return query.order_by(Gasto.creado_en.desc()).all()

def get_ventas_por_metodo(db: Session, req: ExportarReporteRequest):
    query = db.query(
        Venta.metodo_pago,
        func.count(Venta.venta_id).label("cantidad"),
        func.sum(Venta.total).label("total")
    ).filter(Venta.anulada == False)
    if req.fechaInicio:
        query = query.filter(func.date(Venta.creado_en) >= req.fechaInicio)
    if req.fechaFin:
        query = query.filter(func.date(Venta.creado_en) <= req.fechaFin)
    return query.group_by(Venta.metodo_pago).all()

def obtener_periodo_txt(req: ExportarReporteRequest):
    if req.fechaInicio and req.fechaFin:
        if req.fechaInicio == req.fechaFin:
            return f"{req.fechaInicio}"
        return f"{req.fechaInicio} a {req.fechaFin}"
    elif req.fechaInicio:
        return f"Desde {req.fechaInicio}"
    elif req.fechaFin:
        return f"Hasta {req.fechaFin}"
    return "Historico completo"


# ─────────────────────────────────────────
# MATPLOTLIB CHART GENERATORS
# ─────────────────────────────────────────
CAFE_OSCURO = '#2c1810'
CAFE_DORADO = '#d4a96a'
CAFE_CLARO  = '#f5f0eb'
VERDE       = '#16a34a'
ROJO        = '#e11d48'

def _chart_to_image(fig, width_cm=16, height_cm=8):
    """Convert a matplotlib figure to a ReportLab Image flowable."""
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    img = RLImage(buf, width=width_cm*cm, height=height_cm*cm)
    return img

def generar_grafica_financiero(total_ventas, total_gastos, ganancias):
    fig, ax = plt.subplots(figsize=(7, 3.5))
    categorias = ['Ventas\nTotales', 'Gastos\nTotales', 'Ganancia\nNeta']
    valores = [total_ventas, total_gastos, ganancias]
    colores = [VERDE, ROJO, CAFE_DORADO]

    bars = ax.bar(categorias, valores, color=colores, width=0.5, edgecolor='white', linewidth=1.5)
    for bar, val in zip(bars, valores):
        ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max(valores)*0.02,
                f'${val:,.2f}', ha='center', va='bottom', fontsize=9, fontweight='bold', color=CAFE_OSCURO)

    ax.set_title('Rendimiento Financiero', fontsize=13, fontweight='bold', color=CAFE_OSCURO, pad=12)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#d4c4b0')
    ax.spines['bottom'].set_color('#d4c4b0')
    ax.tick_params(colors='#5a3a2a')
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    return _chart_to_image(fig)

def generar_grafica_productos(productos, max_items=10):
    nombres = [p.nombre[:20] for p in productos[:max_items]][::-1]
    cantidades = [int(p.total_vendido or 0) for p in productos[:max_items]][::-1]

    fig, ax = plt.subplots(figsize=(7, max(3, len(nombres)*0.45)))
    bars = ax.barh(nombres, cantidades, color=CAFE_DORADO, edgecolor=CAFE_OSCURO, linewidth=0.5, height=0.6)
    for bar, val in zip(bars, cantidades):
        ax.text(bar.get_width() + max(cantidades)*0.02, bar.get_y() + bar.get_height()/2.,
                str(val), ha='left', va='center', fontsize=8, fontweight='bold', color=CAFE_OSCURO)

    ax.set_title('Productos Más Vendidos', fontsize=13, fontweight='bold', color=CAFE_OSCURO, pad=12)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#d4c4b0')
    ax.spines['bottom'].set_color('#d4c4b0')
    ax.tick_params(colors='#5a3a2a', labelsize=8)
    ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    h = max(3.5, len(nombres)*0.45 + 1.5)
    return _chart_to_image(fig, height_cm=h)

def generar_grafica_metodo_pago(datos_metodo):
    if not datos_metodo:
        return None
    labels = [d.metodo_pago for d in datos_metodo]
    sizes = [float(d.total or 0) for d in datos_metodo]
    colores_pie = ['#2c1810', '#d4a96a', '#8b5e1a', '#a08060', '#5a3a2a']

    fig, ax = plt.subplots(figsize=(5, 3.5))
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, autopct='%1.1f%%', startangle=90,
        colors=colores_pie[:len(labels)],
        textprops={'fontsize': 9, 'color': CAFE_OSCURO},
        wedgeprops={'edgecolor': 'white', 'linewidth': 2}
    )
    for t in autotexts:
        t.set_fontweight('bold')
        t.set_color('white')
        t.set_fontsize(8)
    ax.set_title('Distribución por Método de Pago', fontsize=12, fontweight='bold', color=CAFE_OSCURO, pad=10)
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    return _chart_to_image(fig, width_cm=10, height_cm=7)

def generar_grafica_gastos_cat(gastos):
    if not gastos:
        return None
    cat_totals = {}
    for g in gastos:
        cat_name = g.categoria_gasto.nombre if g.categoria_gasto else "Otros"
        cat_totals[cat_name] = cat_totals.get(cat_name, 0) + float(g.monto or 0)

    if not cat_totals:
        return None

    labels = list(cat_totals.keys())
    valores = list(cat_totals.values())
    colores_gas = ['#e11d48', '#c0392b', '#a93226', '#922b21', '#7b241c']

    fig, ax = plt.subplots(figsize=(6, 3.5))
    bars = ax.bar(labels, valores, color=colores_gas[:len(labels)], width=0.5, edgecolor='white', linewidth=1.5)
    for bar, val in zip(bars, valores):
        ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max(valores)*0.02,
                f'${val:,.2f}', ha='center', va='bottom', fontsize=8, fontweight='bold', color=CAFE_OSCURO)

    ax.set_title('Gastos por Categoría', fontsize=12, fontweight='bold', color=CAFE_OSCURO, pad=12)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#d4c4b0')
    ax.spines['bottom'].set_color('#d4c4b0')
    ax.tick_params(colors='#5a3a2a')
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')
    fig.tight_layout()
    return _chart_to_image(fig, height_cm=7)


# ─────────────────────────────────────────
# EXPORTAR PDF
# ─────────────────────────────────────────
@router.post("/pdf")
def exportar_pdf(
    req: ExportarReporteRequest,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_rol("ADMIN"))
):
    ventas   = get_ventas_filtradas(db, req)
    gastos   = get_gastos_filtrados(db, req)
    productos = get_productos_filtrados(db, req) if req.incluirProductos else []
    metodos  = get_ventas_por_metodo(db, req)

    total_ventas  = sum(float(v.total) for v in ventas)
    total_gastos  = sum(float(g.monto) for g in gastos)
    ganancias     = total_ventas - total_gastos
    periodo_txt   = obtener_periodo_txt(req)

    buffer = BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)

    COLOR_CAFE   = colors.HexColor("#2c1810")
    COLOR_DORADO = colors.HexColor("#8b5e1a")
    COLOR_CLARO  = colors.HexColor("#f5f0eb")

    estilo_titulo = ParagraphStyle("titulo",
        fontSize=20, fontName="Helvetica-Bold",
        textColor=COLOR_CAFE, spaceAfter=4, alignment=TA_LEFT)
    estilo_sub = ParagraphStyle("sub",
        fontSize=10, fontName="Helvetica",
        textColor=COLOR_DORADO, spaceAfter=2)
    estilo_seccion = ParagraphStyle("seccion",
        fontSize=12, fontName="Helvetica-Bold",
        textColor=COLOR_CAFE, spaceBefore=16, spaceAfter=8)
    estilo_normal = ParagraphStyle("normal",
        fontSize=9, fontName="Helvetica", textColor=colors.HexColor("#5a3a2a"))

    elementos = []

    # ── ENCABEZADO
    elementos.append(Paragraph("EspressoPro", estilo_titulo))
    elementos.append(Paragraph(f"Reporte de Ventas — {periodo_txt}", estilo_sub))
    elementos.append(Paragraph(f"Generado el {date.today().strftime('%d/%m/%Y')}   |   Administrador: {admin.nombre} {admin.apellido}", estilo_normal))
    elementos.append(Spacer(1, 0.5*cm))

    # ── RESUMEN GENERAL
    elementos.append(Paragraph("Resumen General", estilo_seccion))
    ticket_promedio = total_ventas / len(ventas) if ventas else 0
    datos_resumen = [
        ["Concepto", "Monto"],
        ["Ventas Totales",      f"${total_ventas:,.2f}"],
        ["Gastos Totales",      f"${total_gastos:,.2f}"],
        ["Ganancias Netas",     f"${ganancias:,.2f}"],
        ["Total Transacciones", str(len(ventas))],
        ["Ticket Promedio",     f"${ticket_promedio:,.2f}"],
    ]
    tabla_resumen = Table(datos_resumen, colWidths=[10*cm, 5*cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), COLOR_CAFE),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 9),
        ("ALIGN",        (1,0), (1,-1), "RIGHT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [COLOR_CLARO, colors.white]),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#d4c4b0")),
        ("PADDING",      (0,0), (-1,-1), 6),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 0.5*cm))

    # ── GRÁFICAS (si están habilitadas)
    if req.incluirGraficas:
        elementos.append(Paragraph("Análisis Gráfico", estilo_seccion))

        # Gráfica 1: Rendimiento Financiero
        img_financiero = generar_grafica_financiero(total_ventas, total_gastos, ganancias)
        elementos.append(img_financiero)
        elementos.append(Spacer(1, 0.4*cm))

        # Gráfica 2: Productos Más Vendidos
        if productos:
            img_productos = generar_grafica_productos(productos)
            elementos.append(img_productos)
            elementos.append(Spacer(1, 0.4*cm))

        # Gráfica 3: Distribución por Método de Pago
        if metodos:
            img_metodo = generar_grafica_metodo_pago(metodos)
            if img_metodo:
                elementos.append(img_metodo)
                elementos.append(Spacer(1, 0.4*cm))

        # Gráfica 4: Gastos por Categoría
        if gastos:
            img_gastos = generar_grafica_gastos_cat(gastos)
            if img_gastos:
                elementos.append(img_gastos)

        elementos.append(PageBreak())

    # ── DETALLE DE VENTAS
    elementos.append(Paragraph("Detalle de Ventas", estilo_seccion))
    if ventas:
        cabecera_ventas = ["ID", "Pedido", "Método Pago", "Subtotal"]
        if req.incluirImpuestos:
            cabecera_ventas.append("IVA")
        cabecera_ventas.append("Total")
        if req.incluirNotas:
            cabecera_ventas.append("Notas")

        datos_ventas = [cabecera_ventas]
        for v in ventas:
            fila = [
                str(v.venta_id),
                f"#{v.pedido_id}",
                v.metodo_pago,
                f"${float(v.subtotal):,.2f}"
            ]
            if req.incluirImpuestos:
                fila.append(f"${float(v.impuesto):,.2f}")
            fila.append(f"${float(v.total):,.2f}")
            if req.incluirNotas:
                observaciones = v.pedido.observaciones if (v.pedido and v.pedido.observaciones) else "N/A"
                fila.append(observaciones[:30])
            datos_ventas.append(fila)

        ancho_col = 16.0 / len(cabecera_ventas)
        tabla_ventas = Table(datos_ventas, colWidths=[ancho_col*cm]*len(cabecera_ventas))
        tabla_ventas.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), COLOR_CAFE),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ALIGN",         (3,0), (-1,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [COLOR_CLARO, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#d4c4b0")),
            ("PADDING",       (0,0), (-1,-1), 5),
        ]))
        elementos.append(tabla_ventas)
    else:
        elementos.append(Paragraph("Sin ventas en este periodo.", estilo_normal))

    elementos.append(Spacer(1, 0.5*cm))

    # ── PRODUCTOS MAS VENDIDOS (tabla)
    if req.incluirProductos and productos:
        elementos.append(Paragraph("Desglose de Productos Vendidos", estilo_seccion))
        datos_prod = [["#", "Producto", "Unidades Vendidas", "Ingresos"]]
        for i, p in enumerate(productos, 1):
            datos_prod.append([
                str(i),
                p.nombre,
                str(int(p.total_vendido or 0)),
                f"${float(p.ingresos or 0):,.2f}",
            ])
        tabla_prod = Table(datos_prod, colWidths=[1*cm, 8*cm, 4*cm, 4*cm])
        tabla_prod.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), COLOR_CAFE),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ALIGN",         (2,0), (-1,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [COLOR_CLARO, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#d4c4b0")),
            ("PADDING",       (0,0), (-1,-1), 5),
        ]))
        elementos.append(tabla_prod)
        elementos.append(Spacer(1, 0.5*cm))

    # ── DETALLE DE GASTOS
    if gastos:
        elementos.append(Paragraph("Detalle de Gastos", estilo_seccion))
        datos_gastos = [["#", "Concepto", "Categoría", "Monto", "Fecha"]]
        for i, g in enumerate(gastos, 1):
            cat_nombre = g.categoria_gasto.nombre if g.categoria_gasto else "Otros"
            datos_gastos.append([
                str(i),
                (g.concepto or "")[:40],
                cat_nombre,
                f"${float(g.monto):,.2f}",
                str(g.fecha_gasto) if g.fecha_gasto else ""
            ])
        tabla_gastos = Table(datos_gastos, colWidths=[1*cm, 6*cm, 3*cm, 3.5*cm, 3*cm])
        tabla_gastos.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#c0392b")),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ALIGN",         (3,0), (3,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [COLOR_CLARO, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#d4c4b0")),
            ("PADDING",       (0,0), (-1,-1), 5),
        ]))
        elementos.append(tabla_gastos)
        elementos.append(Spacer(1, 0.5*cm))

    # ── ESTADO DEL INVENTARIO
    insumos = db.query(Insumo).filter(Insumo.activo == True).all()
    elementos.append(Paragraph("Estado del Inventario", estilo_seccion))
    if insumos:
        datos_insumos = [["Insumo", "Stock Actual", "Stock Mínimo", "Estado", "Costo Unit."]]
        for ins in insumos:
            stock_act = float(ins.stock_actual)
            stock_min = float(ins.stock_minimo)
            if stock_act <= 0:
                estado = "SIN STOCK"
            elif stock_act <= stock_min:
                estado = "BAJO STOCK"
            else:
                estado = "OK"
            datos_insumos.append([
                ins.nombre,
                f"{stock_act:,.2f}",
                f"{stock_min:,.2f}",
                estado,
                f"${float(ins.costo_unitario):,.2f}"
            ])
        tabla_insumos = Table(datos_insumos, colWidths=[5*cm, 3*cm, 3*cm, 2.5*cm, 3*cm])
        tabla_insumos.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), COLOR_CAFE),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ALIGN",         (1,0), (-1,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [COLOR_CLARO, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#d4c4b0")),
            ("PADDING",       (0,0), (-1,-1), 5),
        ]))
        elementos.append(tabla_insumos)
    else:
        elementos.append(Paragraph("Sin insumos registrados.", estilo_normal))

    elementos.append(Spacer(1, 0.5*cm))

    # ── DISTRIBUCIÓN POR MÉTODO DE PAGO (tabla)
    if metodos:
        elementos.append(Paragraph("Resumen por Método de Pago", estilo_seccion))
        datos_met = [["Método", "Transacciones", "Total"]]
        for m in metodos:
            datos_met.append([m.metodo_pago, str(m.cantidad), f"${float(m.total):,.2f}"])
        tabla_met = Table(datos_met, colWidths=[6*cm, 5*cm, 5*cm])
        tabla_met.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), COLOR_CAFE),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("ALIGN",        (1,0), (-1,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [COLOR_CLARO, colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#d4c4b0")),
            ("PADDING",      (0,0), (-1,-1), 6),
        ]))
        elementos.append(tabla_met)

    elementos.append(Spacer(1, 0.5*cm))

    # ── PIE DE PAGINA
    elementos.append(Paragraph(
        "Generado automáticamente por EspressoPro Admin  |  Confidencial",
        ParagraphStyle("pie", fontSize=8, textColor=colors.HexColor("#a08060"),
                       alignment=TA_CENTER)
    ))

    doc.build(elementos)
    buffer.seek(0)

    nombre_archivo = f"Reporte_EspressoPro_{date.today().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


# ─────────────────────────────────────────
# EXPORTAR XLSX
# ─────────────────────────────────────────
@router.post("/xlsx")
def exportar_xlsx(
    req: ExportarReporteRequest,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_rol("ADMIN"))
):
    ventas    = get_ventas_filtradas(db, req)
    gastos    = get_gastos_filtrados(db, req)
    productos = get_productos_filtrados(db, req) if req.incluirProductos else []
    metodos   = get_ventas_por_metodo(db, req)

    total_ventas = sum(float(v.total) for v in ventas)
    total_gastos = sum(float(g.monto) for g in gastos)
    ganancias    = total_ventas - total_gastos
    periodo_txt  = obtener_periodo_txt(req)

    wb = openpyxl.Workbook()

    COLOR_CAFE   = "2c1810"
    COLOR_CLARO  = "f5f0eb"
    COLOR_DORADO = "8b5e1a"
    COLOR_ROJO   = "c0392b"
    thin_border  = Border(
        left=Side(style='thin', color='d4c4b0'),
        right=Side(style='thin', color='d4c4b0'),
        top=Side(style='thin', color='d4c4b0'),
        bottom=Side(style='thin', color='d4c4b0')
    )

    def estilo_header(ws, fila, cols):
        for col in range(1, cols+1):
            cell = ws.cell(row=fila, column=col)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=COLOR_CAFE)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin_border

    def estilo_fila(ws, fila, cols, par=False):
        for col in range(1, cols+1):
            cell = ws.cell(row=fila, column=col)
            cell.border = thin_border
            if par:
                cell.fill = PatternFill("solid", fgColor=COLOR_CLARO)

    # ── HOJA 1: RESUMEN
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1["A1"] = "EspressoPro — Reporte de Ventas"
    ws1["A1"].font = Font(bold=True, size=14, color=COLOR_CAFE)
    ws1["A2"] = f"Periodo: {periodo_txt}  |  Generado: {date.today().strftime('%d/%m/%Y')}  |  Admin: {admin.nombre}"
    ws1["A2"].font = Font(size=10, color=COLOR_DORADO)
    ws1.append([])
    ws1.append(["Concepto", "Monto"])
    estilo_header(ws1, 4, 2)
    ticket_promedio = total_ventas / len(ventas) if ventas else 0
    datos = [
        ("Ventas Totales",      total_ventas),
        ("Gastos Totales",      total_gastos),
        ("Ganancias Netas",     ganancias),
        ("Total Transacciones", len(ventas)),
        ("Ticket Promedio",     ticket_promedio),
    ]
    for i, (concepto, monto) in enumerate(datos, 5):
        ws1.cell(row=i, column=1, value=concepto)
        c2 = ws1.cell(row=i, column=2, value=monto)
        if isinstance(monto, float):
            c2.number_format = '$#,##0.00'
        estilo_fila(ws1, i, 2, par=(i % 2 == 0))
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20

    # ── HOJA 2: VENTAS
    ws2 = wb.create_sheet("Ventas")
    cabecera_ventas = ["ID Venta", "Pedido", "Método Pago", "Subtotal"]
    if req.incluirImpuestos:
        cabecera_ventas.append("IVA")
    cabecera_ventas += ["Descuento", "Total"]
    if req.incluirNotas:
        cabecera_ventas.append("Notas")
    cabecera_ventas.append("Fecha")

    ws2.append(cabecera_ventas)
    estilo_header(ws2, 1, len(cabecera_ventas))
    for i, v in enumerate(ventas, 2):
        fila = [v.venta_id, f"#{v.pedido_id}", v.metodo_pago, float(v.subtotal)]
        if req.incluirImpuestos:
            fila.append(float(v.impuesto))
        fila += [float(v.descuento), float(v.total)]
        if req.incluirNotas:
            fila.append(v.pedido.observaciones if (v.pedido and v.pedido.observaciones) else "")
        fila.append(v.creado_en.strftime("%d/%m/%Y %H:%M") if v.creado_en else "")
        ws2.append(fila)
        estilo_fila(ws2, i, len(cabecera_ventas), par=(i % 2 == 0))

    for col in range(1, len(cabecera_ventas) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ── HOJA 3: PRODUCTOS
    if req.incluirProductos and productos:
        ws3 = wb.create_sheet("Productos")
        ws3.append(["#", "Producto", "Unidades Vendidas", "Ingresos"])
        estilo_header(ws3, 1, 4)
        for i, p in enumerate(productos, 2):
            ws3.append([i-1, p.nombre, int(p.total_vendido or 0), float(p.ingresos or 0)])
            estilo_fila(ws3, i, 4, par=(i % 2 == 0))
            ws3.cell(row=i, column=4).number_format = '$#,##0.00'
        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 30
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 16

        # Gráfica de productos en Excel
        if len(productos) > 0:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = "Productos Más Vendidos"
            chart.y_axis.title = "Unidades"
            data_ref = Reference(ws3, min_col=3, min_row=1, max_row=len(productos)+1)
            cats_ref = Reference(ws3, min_col=2, min_row=2, max_row=len(productos)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 20
            chart.height = 12
            ws3.add_chart(chart, f"F2")

    # ── HOJA 4: GASTOS
    ws4 = wb.create_sheet("Gastos")
    ws4.append(["#", "Concepto", "Categoría", "Monto", "Fecha"])
    estilo_header(ws4, 1, 5)
    for i, g in enumerate(gastos, 2):
        cat_nombre = g.categoria_gasto.nombre if g.categoria_gasto else "Otros"
        ws4.append([i-1, g.concepto or "", cat_nombre, float(g.monto), str(g.fecha_gasto) if g.fecha_gasto else ""])
        estilo_fila(ws4, i, 5, par=(i % 2 == 0))
        ws4.cell(row=i, column=4).number_format = '$#,##0.00'
    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 16
    ws4.column_dimensions["E"].width = 14

    # ── HOJA 5: INVENTARIO
    ws5 = wb.create_sheet("Inventario")
    ws5.append(["Insumo", "Stock Actual", "Stock Mínimo", "Estado", "Costo Unitario"])
    estilo_header(ws5, 1, 5)
    insumos = db.query(Insumo).filter(Insumo.activo == True).all()
    for i, ins in enumerate(insumos, 2):
        stock_act = float(ins.stock_actual)
        stock_min = float(ins.stock_minimo)
        estado = "SIN STOCK" if stock_act <= 0 else ("BAJO STOCK" if stock_act <= stock_min else "OK")
        ws5.append([ins.nombre, stock_act, stock_min, estado, float(ins.costo_unitario)])
        estilo_fila(ws5, i, 5, par=(i % 2 == 0))
        ws5.cell(row=i, column=5).number_format = '$#,##0.00'
        # Color de estado
        estado_cell = ws5.cell(row=i, column=4)
        if estado == "SIN STOCK":
            estado_cell.font = Font(bold=True, color="FFFFFF")
            estado_cell.fill = PatternFill("solid", fgColor="c0392b")
        elif estado == "BAJO STOCK":
            estado_cell.font = Font(bold=True, color="8b5e1a")
            estado_cell.fill = PatternFill("solid", fgColor="fce4b8")
    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 16
    ws5.column_dimensions["C"].width = 16
    ws5.column_dimensions["D"].width = 14
    ws5.column_dimensions["E"].width = 16

    # ── HOJA 6: MÉTODO DE PAGO
    if metodos:
        ws6 = wb.create_sheet("Métodos de Pago")
        ws6.append(["Método", "Transacciones", "Total"])
        estilo_header(ws6, 1, 3)
        for i, m in enumerate(metodos, 2):
            ws6.append([m.metodo_pago, int(m.cantidad), float(m.total)])
            estilo_fila(ws6, i, 3, par=(i % 2 == 0))
            ws6.cell(row=i, column=3).number_format = '$#,##0.00'
        ws6.column_dimensions["A"].width = 20
        ws6.column_dimensions["B"].width = 18
        ws6.column_dimensions["C"].width = 18

        # Gráfica de pie para métodos de pago
        pie = PieChart()
        pie.title = "Distribución por Método de Pago"
        pie.style = 10
        data_ref = Reference(ws6, min_col=3, min_row=1, max_row=len(metodos)+1)
        cats_ref = Reference(ws6, min_col=1, min_row=2, max_row=len(metodos)+1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 16
        pie.height = 12
        ws6.add_chart(pie, "E2")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"Reporte_EspressoPro_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


# ─────────────────────────────────────────
# COMPAT LEGACY (GET endpoints)
# ─────────────────────────────────────────
@router.get("/pdf")
def exportar_pdf_get(
    dias: int = 1,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_rol("ADMIN"))
):
    """Legacy GET endpoint for backward compatibility with Flask export buttons."""
    desde = (date.today() - timedelta(days=dias)).isoformat()
    hoy   = date.today().isoformat()
    req = ExportarReporteRequest(
        fechaInicio=desde, fechaFin=hoy,
        incluirProductos=True, incluirImpuestos=True, incluirGraficas=True
    )
    return exportar_pdf(req, db, admin)

@router.get("/xlsx")
def exportar_xlsx_get(
    dias: int = 1,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(require_rol("ADMIN"))
):
    """Legacy GET endpoint for backward compatibility."""
    desde = (date.today() - timedelta(days=dias)).isoformat()
    hoy   = date.today().isoformat()
    req = ExportarReporteRequest(
        fechaInicio=desde, fechaFin=hoy,
        incluirProductos=True, incluirImpuestos=True, incluirGraficas=True
    )
    return exportar_xlsx(req, db, admin)